import math
import pygame
from datetime import datetime
import time
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
pygame.init()


wb = load_workbook('Book.xlsx')
wb1 = load_workbook('Book5.xlsx')

ws = wb.active
ws1 = wb1.active
print(ws['A1'].value)

print(ws['B60'].value)

width = 800
height = 800
win = pygame.display.set_mode((width,height))
pygame.display.set_caption("clock")
font = pygame.font.SysFont(None, 40)
white = (255,255,255)
black = (0,0,0)
blueblack=(131,145,233)
yellow = (255,255,0)
red = (255,0,0)
clock = pygame.time.Clock()
cpt = win.get_rect().center
angle=0
angle1= 0

angle2=0
#radius = 200

def time1():
    global angle
    ti= datetime.now()
    time2=ti.strftime("%H:%M:%S")#24:60:60
    time3=time.localtime()
    na2 =((time3.tm_sec))
    for i in range(1,60):
        for j in range(1,3):
            char = get_column_letter(j)
            for r in range(1,2):
                char2 = get_column_letter(r)
            if na2 == ws[char+str(i)].value:
                angle = ws[char2+str(i)].value          
                
          
    draw = font.render(time2,True,black)
    win.blit(draw, draw.get_rect(center = cpt))
    draw1 = font.render(str(12),True,black)
    win.blit(draw1,(390,210))
    draw2 = font.render(str(1),True,black)
    win.blit(draw2,(490,235))
    draw3 = font.render(str(2),True,black)
    win.blit(draw3,(552,300))
    draw4 = font.render(str(3),True,black)
    win.blit(draw4,(575,390))
    draw5 = font.render(str(4),True,black)
    win.blit(draw5,(552,480))
    draw6 = font.render(str(5),True,black)
    win.blit(draw6,(490,545))
    draw7 = font.render(str(6),True,black)
    win.blit(draw7,(392.5,570))
    draw8 = font.render(str(7),True,black)
    win.blit(draw8,(292.5,545))
    draw9 = font.render(str(8),True,black)
    win.blit(draw9,(230,478))
    draw10 = font.render(str(9),True,black)
    win.blit(draw10,(210,390))
    draw11 = font.render(str(10),True,black)
    win.blit(draw11,(230,300))
    draw12 = font.render(str(11),True,black)
    win.blit(draw12,(290,235))
    
                            
        
    
    
    pygame.draw.circle(win,black,(400,400),210,10)
    pygame.draw.line(win, yellow, cpt, (pt_x, pt_y), 5)
    #pygame.draw.line(win, red, cpt, (pt_x1, pt_y1), 2)
    #pygame.draw.line(win, blueblack, cpt, (pt_x2, pt_y2), 2)
    
    return angle
    #return angle1
    #return angle2


def time2():
    global angle1
    time4=time.localtime()
    
    na1=((time4.tm_min*60)+(time4.tm_sec))
    
    for m in range(1,3600):
        for n in range(3,5):
            char = get_column_letter(n)
            for o in range(3,4):
                char2 = get_column_letter(o)
            if na1 == ws[char+str(m)].value:
                angle1 = ws[char2+str(m)].value
    pygame.draw.line(win, red, cpt, (pt_x1, pt_y1), 5)           
    return angle1


def time3():
    global angle2
    time5=time.localtime()
    #print(time5)
    na=((time5.tm_hour*60*60)+(time5.tm_min*60)+(time5.tm_sec))
    print(na)
    for p in range(1,90200):
        for q in range(5,7):
            char = get_column_letter(q)
            for s in range(5,6):
                char2 = get_column_letter(s)
            if na == ws1[char+str(p)].value:
                angle2 = ws1[char2+str(p)].value
    pygame.draw.line(win, blueblack, cpt, (pt_x2, pt_y2), 5)         
    return angle2
    
run = True
while run:
    clock.tick(1)
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            run = False

    #hands = (yellow,angle1,200)
    #handh = (black,angle2,100)
    ##handm = (red,angle3,150)
    #hand = [hands,handh,handm]
       
    angle += 6
    if angle >= 360:
          angle = 0
    angle_rad = math.radians(angle)
    pt_x = cpt[0] + 200 * math.sin(angle_rad)
    pt_y = cpt[1] - 200 * math.cos(angle_rad)


    angle1 += 0.1
    if angle1 >= 360:
        angle1= 0
    angle1_rad = math.radians(angle1)
    pt_x1 = cpt[0] + 150 * math.sin(angle1_rad)
    pt_y1 = cpt[1] - 150 * math.cos(angle1_rad)


    angle2 += 0.00833
    if angle2 >= 360:
        angle2 = 0 
    angle2_rad = math.radians(angle2)
    pt_x2 = cpt[0] + 100 * math.sin(angle2_rad)
    pt_y2 = cpt[1] - 100 * math.cos(angle2_rad)

 
    win.fill(white)

    time1()
    time2()
    time3()

    pygame.display.update()
pygame.quit()
